import re

import openpyxl
from openpyxl.styles import Font
from _xhr_tool._utils.excelTool.csvFileOptionTool import csvFileTool
class ExcelTool:
    def __init__(self):
        pass
    def _createNewWorkbook(self):
        return openpyxl.Workbook()
    def _createNewSheet(self,workbook,sheetName):
        return workbook.create_sheet(sheetName)
    def _changeCellToStr(self,cell):
        if cell.hyperlink:
            cell.value="<href='"+cell.hyperlink.target+"'>"+cell.value
        return cell.value
    def _changeStrToCell(self,cell):
        attrss=re.findall(r'<(.*)>',cell.value)
        if len(attrss)>0:
            arr=[]
            attrs=attrss[0].split(' ')
            for attr in attrs:
                if attr.strip()!='':
                    arr.append(attr)
            for ar in arr:
                try:
                    tu=ar.split('=')
                    if tu[1].startswith("'") and tu[1].endswith("'"):
                        pass
                    elif tu[1].startswith("\"") and tu[1].endswith("\""):
                        pass
                    else:
                        raise ValueError()
                except:
                    raise ValueError("语法错误"+ar+"：无法处理该语句，请确保属性中有等号")
                if tu[0]=='href':############################################################################此处扩展
                    cell.hyperlink=tu[1][1:len(tu[1])-1]
                    cell.value=re.sub(r'<.*>','',cell.value)
                    cell.font=Font(bold=False,italic=False,underline="single",color='0000FF')
                else:
                    raise ValueError('语法错误：不存在'+tu[0]+'这个属性,目前只允许以下属性存在：href')


        return cell.value
    def _changeStyle(self,cell,fontStyle={}):
        """
        改变文字的样式fontStype
        underline:single | none _添加下划线
        fontColor: 0000FF _蓝色
        size: 11
        bold: True _是否加粗
        name: 宋体 _字体样式
        italic: True   _是否斜体
        strike: False _是否添加删除线
        :param cell:
        :param style:
        :return:
        """
        #设置默认样式
        if "size" in fontStyle.keys():
            pass
        cell.font = Font(size=fontStyle['size']|100, bold=True, italic=True, underline="single", color='0000FF')
        return cell
    def optionExecl(self,path='',sheetName='新建表格',datas=[],mode="",styleRemain=True):
        """
        一体化
        :param path:
        :param sheetName:
        :param datas:
        :param mode:
        :param styleRemain:
        :return:
        """
        if mode=='w':
            wb=self._createNewWorkbook()
            sheet=self._createNewSheet(wb,sheetName)
            #表头的写入

            firstArr=list(datas[0].keys())
            for i in range(0,len(firstArr)):
                sheet.cell(row=1,column=i+1,value=firstArr[i])
            #表体的写入
            for j in range(0,len(datas)):
                for i in range(0,len(firstArr)):
                    cell=sheet.cell(row=j + 2, column=i + 1, value=datas[j][firstArr[i]])
                    #对value进行处理
                    if type(cell.value)==str:
                        cell=self._changeStrToCell(cell)
            del wb['Sheet']#删除默认表单
            wb.save(path)
        elif mode=='r':
            wb = openpyxl.load_workbook(path)
            sh = wb[sheetName]
            rows_data = list(sh.rows)
            headLine=[]#头
            datas=[]
            for hc in rows_data.pop(0):
                headLine.append(hc.value)
            for row in rows_data:
                data={}
                for i in range(0,len(row)):
                    if styleRemain:
                        data.setdefault(headLine[i], self._changeCellToStr(row[i]))
                    else:
                        data.setdefault(headLine[i],row[i].value)
                datas.append(data)
            return datas
        elif mode=='a':
            oldDatas=self.optionExecl(path=path, sheetName=sheetName, mode="r", styleRemain=True)
            newDatas=oldDatas+datas
            self.optionExecl(path=path, sheetName=sheetName,datas=newDatas, mode="w")
    def chageCsvToExcelFile(self,path="",encoding='utf-8',sheetName="新建表格"):
        """
        #将csv文件转成excel文件
        :param path:
        :param encoding:
        :param sheetName:
        :return:
        """
        csvTool=csvFileTool()
        csvArr=csvTool.readCsvData_arrDict(path,encoding=encoding)

        self.optionExecl(path=path.replace('.csv','.xlsx'),sheetName=sheetName,mode='w',datas=csvArr)
    def filter(self,path="",attr="",conditionFunction="",sheetName="新建表格"):
        """
        根据条件过滤得到新的数据。

        :param path:
        :param attr:
        :param conditionFunction:
        :param sheetName:
        :return:
        """
        filterItems=[]
        items=self.optionExecl(path=path,sheetName=sheetName,mode='r')
        for item in items:
            if conditionFunction(item[attr]):
                filterItems.append(item)
        return filterItems
if __name__ == '__main__':
    tool=ExcelTool()
    data = tool.optionExecl(path='test.xlsx', sheetName='高强涤纶_顺企网', mode='r')
    #取出excel数据
    #保存成csv数据
    csvtool=csvFileTool()
    csvtool.writeCsvData_arrDict(path='test.csv',arr=data,encoding='utf-8')


    # datas=[{'name':'张三','age':10,'资料':"<href='C:/Users/1234567/Desktop/hrefsFile/舒服的耳语.txt'>舒服的耳语.txt"},{'name': '历史', 'age': 12, '资料': 'hrefsFile/舒服的耳语.txt'}]
    #
    # tool.optionExecl(path='resouse/cases.xlsx',sheetName='112',datas=datas,mode='w')
    # data = tool.optionExecl(path='resouse/cases.xlsx',datas=datas,sheetName='112', mode='a')
    # # data=tool.optionExecl(path='resouse/cases.xlsx',datas=datas ,sheetName='112',mode='w')
